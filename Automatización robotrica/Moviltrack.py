# -*- coding: latin-1 -*-
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from time import sleep
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.common.exceptions import NoSuchElementException
from os import remove
import pandas as pd
import urllib.request
import shutil
import time
import os
import subprocess
import pyautogui as py
from openpyxl import workbook
from openpyxl import load_workbook
from dotenv import load_dotenv
import warnings
warnings.simplefilter("ignore")

op = webdriver.ChromeOptions()
#op.binary_location = "ruta del ejecutable de Chrome"
op.add_argument("--headless")
op.add_argument("--disable-gpu")
op.add_experimental_option('excludeSwitches', ['enable-logging'])
#driver = webdriver.Chrome(chrome_options=op)
#driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
driver = webdriver.Chrome(executable_path='C:/Users/ROBOTRPASI/Desktop/chromedriver.exe')

driver.maximize_window()

load_dotenv()
input_email_id = os.getenv('NAME1')
input_pwd = os.getenv('PASSWORD1')

def executeBot():

        # Get the page. 
        driver.get('http://10.181.3.77:9502/xmlpserver/login.jsp')
        time.sleep(5)
        print("SAC ESTA ABIERTO...")
        
        # We find the xpath for insert our email address.
        email = driver.find_element(By.XPATH,"//*[@id='id']")
        # Put mail.
        email.send_keys(input_email_id)
        print("Email enviado.")
        # We find the xpath for insert our password.
        time.sleep(2)
        # We find xpath of button Login.      
        password = driver.find_element(By.XPATH,"//*[@id='passwd']")
        # Put password.
        password.send_keys(input_pwd)
        print("Password enviado.")
        # We find xpath of button Login.
        button = driver.find_element(By.XPATH,"//*[@id='content-cell']/div/span/table/tbody/tr[2]/td[2]/div/table/tbody/tr[3]/td/form[1]/input[4]")
        # We click on button of login.
        button.click()
        # Open session.
        print("SAC esta abierto.")
        x = 0
        time.sleep(5)
        
        # View Issues
        time.sleep(5)
        driver.get('http://10.181.3.77:9502/xmlpserver/Oracle+Identity+Manager/Usuarios+MovilTrack.xdo')
        # Download report
        time.sleep(5)
        button = driver.find_element(By.XPATH,"//*[@id='xdo:viewFormatIcon']")
        button.click()
        time.sleep(10)
        
        button = driver.find_element(By.XPATH,"//*[@id='_xdoFMenu5']/div/div/ul/li[6]/div/a/div[2]")
        button.click()
        time.sleep(15)
        
        
        #Mover Archivo
        shutil.move("C:/Users/ROBOTRPASI/Downloads/Usuarios MovilTrack_Usuarios MovilTrack.xlsx", "C:/Users/ROBOTRPASI/Pictures/Carpeta de Planos/TMP_MOVILTRACK_USUARIOS.xlsx")
        
    
        driver.close()
        #proc.terminate()
executeBot()
wb = load_workbook("C:/Users/ROBOTRPASI/Pictures/Carpeta de Planos/TMP_MOVILTRACK_USUARIOS.xlsx", read_only=False)


ws=wb["Sheet1"]

ws.unmerge_cells("A1:F1")
ws.delete_rows(1)


wb.save("C:/Users/ROBOTRPASI/Pictures/Carpeta de Planos/TMP_MOVILTRACK_USUARIOS.xlsx")
time.sleep(5)

archivo_excel = 'C:/Users/ROBOTRPASI/Pictures/Carpeta de Planos/TMP_MOVILTRACK_USUARIOS.xlsx'

df = pd.read_excel(archivo_excel)

archivo_txt = 'S:/PLANOS/TMP_MOVILTRACK_USUARIOS.txt'

df.to_csv(archivo_txt, sep='\t', index=False)


print('Terminado')