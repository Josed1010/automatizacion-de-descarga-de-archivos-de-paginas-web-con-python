import pandas as pd
import glob
import time
import openpyxl
from tkinter import filedialog
import warnings
warnings.simplefilter("ignore")
import pyautogui as py
from openpyxl import workbook
from openpyxl import load_workbook





folder_selected = 'C:/Users/ROBOTRPASI/Pictures/Planos SIGA'

if not folder_selected:
    exit()
print(folder_selected)

file_list = glob.glob(folder_selected + "/*.xlsx" )

print(file_list)

excel_list = []
for file in file_list:
    excel_list.append(pd.read_excel(file))
    

excel_merged = pd.DataFrame()

for excel_file in excel_list:
    excel_merged = pd.concat([excel_merged, pd.DataFrame.from_records(excel_file)], ignore_index= False)



excel_merged.to_excel("C:/Users/ROBOTRPASI/Pictures/Carpeta de Planos/TMP_SIGA_USUARIOS.xlsx")


wb = load_workbook("C:/Users/ROBOTRPASI/Pictures/Carpeta de Planos/TMP_SIGA_USUARIOS.xlsx", read_only=False)

'''
wb.sheetnames
'''
ws=wb["Sheet1"]


ws.delete_cols(1)
time.sleep(5)


wb.save("C:/Users/ROBOTRPASI/Pictures/Carpeta de Planos/TMP_SIGA_USUARIOS.xlsx")


archivo_excel = 'C:/Users/ROBOTRPASI/Pictures/Carpeta de Planos/TMP_SIGA_USUARIOS.xlsx'

df = pd.read_excel(archivo_excel)

archivo_txt = 'S:/PLANOS/TMP_SIGA_USUARIOS.txt'

df.to_csv(archivo_txt, sep='\t', index=False)


print('Terminado')