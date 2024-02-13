from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from time import sleep
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.common.exceptions import NoSuchElementException
from os import remove
import pandas as pd
import urllib.request
import shutil
import time
import os
import subprocess
import pyautogui as py
from openpyxl import workbook
from openpyxl import load_workbook
from dotenv import load_dotenv
import warnings
warnings.simplefilter("ignore")

op = webdriver.ChromeOptions()
#op.binary_location = "ruta del ejecutable de Chrome"
op.add_argument("--headless")
op.add_argument("--disable-gpu")
op.add_experimental_option('excludeSwitches', ['enable-logging'])
#driver = webdriver.Chrome(chrome_options=op)
#driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
driver = webdriver.Chrome(executable_path=r'C:/Users/ROBOTRPASI/Desktop/chromedriver.exe')

driver.maximize_window()

load_dotenv()
input_email_id = os.getenv('NAME3')
input_pwd = os.getenv('PASSWORD3')

def executeBot():

        # Get the page. 
        driver.get('http://10.181.0.116:8080/geopayment-jsf-ui/index.xhtml')
        time.sleep(5)
        print("SAC ESTA ABIERTO...")
        # We find the xpath for insert our email address.
        email = driver.find_element(By.XPATH,"//*[@id='j_username']")
        # Put mail.
        email.send_keys(input_email_id)
        print("Email enviado.")
        # We find the xpath for insert our password.
        time.sleep(2)
        # We find xpath of button Login.               
        password = driver.find_element(By.XPATH,"//*[@id='j_password']")
        # Put password.
        password.send_keys(input_pwd)
        print("Password enviado.")
        # We find xpath of button Login.
        button = driver.find_element(By.XPATH,"//*[@id='loginButton']")
        # We click on button of login.
        button.click()
        # Open session.
        print("SAC esta abierto.")
        x = 0
        # View Issues
        time.sleep(2)

        #Botton usuarios
        button = driver.find_element(By.XPATH,"//*[@id='dropMenuForm:j_id14_span']")
        button.click()
        time.sleep(4)
        #boton reoporte de usuarios
        button = driver.find_element(By.XPATH,"//*[@id='dropMenuForm:j_id15:anchor']")
        button.click()
        time.sleep(4)
        
        
        #desplegar el formulario
        button = driver.find_element(By.XPATH,"//*[@id='dropMenuForm:j_id17:anchor']")
        button.click()
        time.sleep(4) 
               
        
        #Mover Archivo
        shutil.move("C:/Users/ROBOTRPASI/Downloads/reporte.xlsx", "C:/Users/ROBOTRPASI/Pictures/Carpeta de Planos/TMP_GEOPAYMENT_USUARIOS.xlsx")
        
        driver.close()
        #proc.terminate()
executeBot()

#encontrar el archivo
wb = load_workbook("C:/Users/ROBOTRPASI/Pictures/Carpeta de Planos/TMP_GEOPAYMENT_USUARIOS.xlsx", read_only=False)

#Encontrar la hoja a modificar
ws=wb["table"]

#Hacer las modificaciones
ws.unmerge_cells("A1:E1")
ws.delete_rows(1)

#Guardar nuevamente
wb.save("C:/Users/ROBOTRPASI/Pictures/Carpeta de Planos/TMP_GEOPAYMENT_USUARIOS.xlsx")
time.sleep(4) 


archivo_excel = 'C:/Users/ROBOTRPASI/Pictures/Carpeta de Planos/TMP_GEOPAYMENT_USUARIOS.xlsx'

df = pd.read_excel(archivo_excel)

archivo_txt = 'S:/PLANOS/TMP_GEOPAYMENT_USUARIOS.txt'

df.to_csv(archivo_txt, sep='\t', index=False)


print('Terminado')