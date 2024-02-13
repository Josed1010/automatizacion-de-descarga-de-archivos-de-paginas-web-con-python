# -*- coding: latin-1 -*-
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from time import sleep
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.common.exceptions import NoSuchElementException
from os import remove
import urllib.request
import shutil
import time
import os
import subprocess
import pyautogui as py
from openpyxl import workbook
from openpyxl import load_workbook
from dotenv import load_dotenv
import warnings
warnings.simplefilter("ignore")

op = webdriver.ChromeOptions()
#op.binary_location = "ruta del ejecutable de Chrome"
op.add_argument("--headless")
op.add_argument("--disable-gpu")
op.add_experimental_option('excludeSwitches', ['enable-logging'])
#driver = webdriver.Chrome(chrome_options=op)
#driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
driver = webdriver.Chrome(executable_path='C:/Users/ROBOTRPASI/Desktop/chromedriver.exe')

driver.maximize_window()

load_dotenv()
input_email_id = os.getenv('NAME')
input_pwd = os.getenv('PASSWORD')

def executeBot():

        # Get the page. 
        driver.get('https://siga.corbeta.com.co/Vistas/Seguridad/IniciarSesion?ReturnUrl=%2FDefault')
        time.sleep(5)
        print("SAC ESTA ABIERTO...")
        
        # We find the xpath for insert our email address.
        email = driver.find_element(By.XPATH,"//*[@id='tbxIdentificacion']")
        # Put mail.
        email.send_keys(input_email_id)
        print("Email enviado.")
        # We find the xpath for insert our password.
        time.sleep(2)
        # We find xpath of button Login.    
        password = driver.find_element(By.XPATH,"//*[@id='tbxClave']")
        # Put password.
        password.send_keys(input_pwd)
        print("Password enviado.")
        # We find xpath of button Login.
        button = driver.find_element(By.XPATH,"//*[@id='btnAutenticar']")
        # We click on button of login.
        button.click()
        # Open session.
        print("SAC esta abierto.")
        x = 0
        # View Issues
        time.sleep(5)
        
        # Download report
        time.sleep(5)
        button = driver.find_element(By.XPATH,"//*[@id='menu']/ul/li[2]/a/i")
        button.click()
        time.sleep(5)
        
        button = driver.find_element(By.XPATH,"//*[@id='menu']/ul/li[2]/ul/li/a/span")
        button.click()
        time.sleep(8)
        
        button = driver.find_element(By.XPATH,"//*[@id='divContenedorBloqueoAccesoHabilitado']/span/span[1]/div/ul/li/input")
        button.click()
        time.sleep(5)
        
        button = driver.find_element(By.XPATH,"//*[@id='divContenedorBloqueoAccesoHabilitado']/span/span[1]/div/div/ul/li[2]")
        button.click()
        time.sleep(5)
        
        button = driver.find_element(By.XPATH,"//*[@id='btnVerReporte']")
        button.click()
        time.sleep(12)
        
        button = driver.find_element(By.XPATH,"//*[@id='ctl00_cprPrincipal_rvrSIGA_ctl09_ctl04_ctl00_ButtonImgDown']")
        button.click()
        time.sleep(12)
         
        button = driver.find_element(By.XPATH,"//*[@id='ctl00_cprPrincipal_rvrSIGA_ctl09_ctl04_ctl00_Menu']/div[2]/a")
        button.click()
        time.sleep(15)
        
        #Mover Archivo
        shutil.move("C:/Users/ROBOTRPASI/Downloads/Reporte.Seguridad.CuentaUsuario_Perfilamiento.xlsx", "C:/Users/ROBOTRPASI/Pictures/Planos SIGA/TMP_SIGA_SI.xlsx")
                
        driver.close()
        #proc.terminate()
executeBot()

wb = load_workbook("C:/Users/ROBOTRPASI/Pictures/Planos SIGA/TMP_SIGA_SI.xlsx", read_only=False)

print(wb.get_sheet_names())
std=wb.get_sheet_by_name('CuentaUsuario - Entes')
wb.remove_sheet(std)

ws=wb["CuentaUsuario - Perfiles"]

ws.unmerge_cells("A1:O1")
ws.delete_rows(1)
ws.delete_rows(2)
ws.delete_rows(3)
ws.unmerge_cells("A4:Q4")
ws.delete_rows(4)
ws.unmerge_cells("A5:Q5")
ws.delete_rows(5)
ws.delete_rows(6)
ws.delete_rows(7)
time.sleep(5)
ws.delete_rows(1)
ws.delete_rows(2)
ws.delete_rows(3)
time.sleep(5)
ws.delete_rows(1)

wb.save("C:/Users/ROBOTRPASI/Pictures/Planos SIGA/TMP_SIGA_SI.xlsx")

print('terminado')