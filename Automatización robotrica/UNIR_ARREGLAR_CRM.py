from time import sleep
import pandas as pd
import urllib.request
import shutil
import time
import os
import subprocess
import pandas as pd
import glob
import pyautogui as py
from openpyxl import workbook
from openpyxl import load_workbook
import warnings
warnings.simplefilter("ignore")



#CRM_AKT_CARTERA
wb = load_workbook("C:/Users/ROBOTRPASI/Pictures/Planos CRM/CRM_AKT_CARTERA.xlsx", read_only=False)
ws=wb["Sheet1"]
ws.unmerge_cells("A1:H1")
ws.delete_rows(1)
wb.save("C:/Users/ROBOTRPASI/Pictures/Planos CRM/CRM_AKT_CARTERA.xlsx")
print("Terminado CRM_AKT_CARTERA")
time.sleep(3)

#CRM_AKT_PQR
wb = load_workbook("C:/Users/ROBOTRPASI/Pictures/Planos CRM/CRM_AKT_PQR.xlsx", read_only=False)
ws=wb["Sheet1"]
ws.unmerge_cells("A1:H1")
ws.delete_rows(1)
wb.save("C:/Users/ROBOTRPASI/Pictures/Planos CRM/CRM_AKT_PQR.xlsx")
print("Terminado CRM_AKT_PQR")
time.sleep(3)

#CRM_AKT_PREVENTA
wb = load_workbook("C:/Users/ROBOTRPASI/Pictures/Planos CRM/CRM_AKT_PREVENTA.xlsx", read_only=False)
ws=wb["Sheet1"]
ws.unmerge_cells("A1:H1")
ws.delete_rows(1)
wb.save("C:/Users/ROBOTRPASI/Pictures/Planos CRM/CRM_AKT_PREVENTA.xlsx")
print("Terminado CRM_AKT_PREVENTA")
time.sleep(3)

#CRM_ALKOMPRAR
wb = load_workbook("C:/Users/ROBOTRPASI/Pictures/Planos CRM/CRM_ALKOMPRAR.xlsx", read_only=False)
ws=wb["Sheet1"]
ws.unmerge_cells("A1:H1")
ws.delete_rows(1)
wb.save("C:/Users/ROBOTRPASI/Pictures/Planos CRM/CRM_ALKOMPRAR.xlsx")
print("Terminado CRM_ALKOMPRAR")
time.sleep(3)

#CRM_ALKOSTO
wb = load_workbook("C:/Users/ROBOTRPASI/Pictures/Planos CRM/CRM_ALKOSTO.xlsx", read_only=False)
ws=wb["Sheet1"]
ws.unmerge_cells("A1:H1")
ws.delete_rows(1)
wb.save("C:/Users/ROBOTRPASI/Pictures/Planos CRM/CRM_ALKOSTO.xlsx")
print("Terminado CRM_ALKOSTO")
time.sleep(3)

#CRM_DISTRIBUCIONES
wb = load_workbook("C:/Users/ROBOTRPASI/Pictures/Planos CRM/CRM_DISTRIBUCIONES.xlsx", read_only=False)
ws=wb["Sheet1"]
ws.unmerge_cells("A1:H1")
ws.delete_rows(1)
wb.save("C:/Users/ROBOTRPASI/Pictures/Planos CRM/CRM_DISTRIBUCIONES.xlsx")
print("Terminado CRM_DISTRIBUCIONES")
time.sleep(3)

#CRM_FOTON_PREVENTA
wb = load_workbook("C:/Users/ROBOTRPASI/Pictures/Planos CRM/CRM_FOTON_PREVENTA.xlsx", read_only=False)
ws=wb["Sheet1"]
ws.unmerge_cells("A1:H1")
ws.delete_rows(1)
wb.save("C:/Users/ROBOTRPASI/Pictures/Planos CRM/CRM_FOTON_PREVENTA.xlsx")
print("Terminado CRM_FOTON_PREVENTA")
time.sleep(3)

#CRM_FOTON_PQR
wb = load_workbook("C:/Users/ROBOTRPASI/Pictures/Planos CRM/CRM_FOTON_PQR.xlsx", read_only=False)
ws=wb["Sheet1"]
ws.unmerge_cells("A1:H1")
ws.delete_rows(1)
wb.save("C:/Users/ROBOTRPASI/Pictures/Planos CRM/CRM_FOTON_PQR.xlsx")
print("Terminado CRM_FOTON_PQR")
time.sleep(3)

#CRM_KALLEY
wb = load_workbook("C:/Users/ROBOTRPASI/Pictures/Planos CRM/CRM_KALLEY.xlsx", read_only=False)
ws=wb["Sheet1"]
ws.unmerge_cells("A1:H1")
ws.delete_rows(1)
wb.save("C:/Users/ROBOTRPASI/Pictures/Planos CRM/CRM_KALLEY.xlsx")
print("Terminado CRM_KALLEY")
time.sleep(3)

#CRM_NARIÑO
wb = load_workbook("C:/Users/ROBOTRPASI/Pictures/Planos CRM/CRM_NARINO.xlsx", read_only=False)
ws=wb["Sheet1"]
ws.unmerge_cells("A1:H1")
ws.delete_rows(1)
wb.save("C:/Users/ROBOTRPASI/Pictures/Planos CRM/CRM_NARINO.xlsx")
print("Terminado CRM_NARINO")
time.sleep(3)

#CRM_PASTO
wb = load_workbook("C:/Users/ROBOTRPASI/Pictures/Planos CRM/CRM_PASTO.xlsx", read_only=False)
ws=wb["Sheet1"]
ws.unmerge_cells("A1:H1")
ws.delete_rows(1)
wb.save("C:/Users/ROBOTRPASI/Pictures/Planos CRM/CRM_PASTO.xlsx")
print("Terminado CRM_PASTO")
time.sleep(3)

#CRM_TEXTILES
wb = load_workbook("C:/Users/ROBOTRPASI/Pictures/Planos CRM/CRM_TEXTILES.xlsx", read_only=False)
ws=wb["Sheet1"]
ws.unmerge_cells("A1:H1")
ws.delete_rows(1)
wb.save("C:/Users/ROBOTRPASI/Pictures/Planos CRM/CRM_TEXTILES.xlsx")
print("Terminado CRM_TEXTILES")
time.sleep(3)

#UNIR PLANOS
folder_selected = 'C:/Users/ROBOTRPASI/Pictures/Planos CRM/'
if not folder_selected:
    exit()
print(folder_selected)

file_list = glob.glob(folder_selected + "/*.xlsx" )

print(file_list)

excel_list = []
for file in file_list:
    excel_list.append(pd.read_excel(file))
    
excel_merged = pd.DataFrame()

for excel_file in excel_list:
    excel_merged = pd.concat([excel_merged, pd.DataFrame.from_records(excel_file)], ignore_index= False)
    
excel_merged.to_excel("C:/Users/ROBOTRPASI/Pictures/Carpeta de Planos/TMP_CRM_USUARIOS.xlsx")

#ELIMINAR LA COLUMNA
wb = load_workbook("C:/Users/ROBOTRPASI/Pictures/Carpeta de Planos/TMP_CRM_USUARIOS.xlsx", read_only=False)
ws=wb["Sheet1"]
ws.delete_cols(1)
time.sleep(3)
ws.delete_cols(7)
time.sleep(3)
ws.delete_cols(7)
wb.save("C:/Users/ROBOTRPASI/Pictures/Carpeta de Planos/TMP_CRM_USUARIOS.xlsx")


#GUARDAR EN LA S
archivo_excel = 'C:/Users/ROBOTRPASI/Pictures/Carpeta de Planos/TMP_CRM_USUARIOS.xlsx'
df = pd.read_excel(archivo_excel)
archivo_txt = 'S:/PLANOS/TMP_CRM_USUARIOS.txt'
df.to_csv(archivo_txt, sep='\t', index=False)

print('Terminado')

